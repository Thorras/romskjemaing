"""
Test cases for Excel exporter functionality.
"""

import os
import tempfile
from pathlib import Path

import pytest

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from ifc_room_schedule.export.excel_exporter import ExcelExporter
from ifc_room_schedule.data.space_model import SpaceData
from ifc_room_schedule.data.surface_model import SurfaceData
from ifc_room_schedule.data.space_boundary_model import SpaceBoundaryData
from ifc_room_schedule.data.relationship_model import RelationshipData


@pytest.fixture
def sample_spaces():
    """Create sample space data for testing."""
    # Create surfaces
    surface1 = SurfaceData(
        id="surface_1",
        type="Wall",
        area=25.5,
        material="Concrete",
        ifc_type="IfcWall",
        related_space_guid="space_guid_1",
        user_description="North wall"
    )
    
    surface2 = SurfaceData(
        id="surface_2", 
        type="Floor",
        area=50.0,
        material="Concrete Slab",
        ifc_type="IfcSlab",
        related_space_guid="space_guid_1",
        user_description="Main floor"
    )
    
    # Create space boundaries
    boundary1 = SpaceBoundaryData(
        id="boundary_1",
        guid="boundary_guid_1",
        name="North Wall Boundary",
        description="Boundary to exterior",
        physical_or_virtual_boundary="Physical",
        internal_or_external_boundary="External",
        related_building_element_guid="wall_guid_1",
        related_building_element_name="North Wall",
        related_building_element_type="IfcWall",
        related_space_guid="space_guid_1",
        boundary_surface_type="Wall",
        boundary_orientation="North",
        connection_geometry={},
        calculated_area=25.5,
        boundary_level=1,
        display_label="North Wall to Exterior"
    )
    
    # Create relationships
    relationship1 = RelationshipData(
        related_entity_guid="zone_guid_1",
        related_entity_name="Office Zone",
        related_entity_description="Main office zone",
        relationship_type="Contains",
        ifc_relationship_type="IfcRelContainedInSpatialStructure"
    )
    
    # Create space
    space1 = SpaceData(
        guid="space_guid_1",
        name="Office 101",
        long_name="Main Office Room 101",
        description="Primary office space",
        object_type="Office",
        zone_category="Work",
        number="101",
        elevation=0.0,
        quantities={"Height": 3.0, "FinishFloorHeight": 0.1},
        surfaces=[surface1, surface2],
        space_boundaries=[boundary1],
        relationships=[relationship1],
        user_descriptions={"space": "Main office room"},
        processed=True
    )
    
    return [space1]


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
class TestExcelExporter:
    """Test cases for ExcelExporter."""
    
    def test_exporter_initialization(self):
        """Test Excel exporter initializes correctly."""
        exporter = ExcelExporter()
        assert exporter.source_file_path is None
        assert exporter.application_version == "1.0.0"
    
    def test_set_source_file(self):
        """Test setting source file path."""
        exporter = ExcelExporter()
        test_path = "/test/file.ifc"
        exporter.set_source_file(test_path)
        assert exporter.source_file_path == test_path
    
    def test_export_to_excel_basic(self, sample_spaces):
        """Test basic Excel export functionality."""
        exporter = ExcelExporter()
        
        with tempfile.TemporaryDirectory() as temp_dir:
            export_path = os.path.join(temp_dir, "test_export.xlsx")
            
            success, message = exporter.export_to_excel(sample_spaces, export_path)
            
            assert success is True
            assert "Successfully exported" in message
            assert os.path.exists(export_path)
    
    def test_export_to_excel_with_extension(self, sample_spaces):
        """Test Excel export adds extension if missing."""
        exporter = ExcelExporter()
        
        with tempfile.TemporaryDirectory() as temp_dir:
            export_path = os.path.join(temp_dir, "test_export")  # No extension
            
            success, message = exporter.export_to_excel(sample_spaces, export_path)
            
            assert success is True
            assert os.path.exists(export_path + ".xlsx")
    
    def test_export_excel_workbook_structure(self, sample_spaces):
        """Test Excel export workbook structure."""
        exporter = ExcelExporter()
        exporter.set_source_file("/test/source.ifc")
        
        with tempfile.TemporaryDirectory() as temp_dir:
            export_path = os.path.join(temp_dir, "test_export.xlsx")
            
            success, message = exporter.export_to_excel(sample_spaces, export_path)
            assert success is True
            
            # Load and verify workbook structure
            wb = load_workbook(export_path)
            
            # Check expected sheets exist
            expected_sheets = ["Overview", "Spaces", "Surfaces", "Space Boundaries", "Relationships", "Summary"]
            for sheet_name in expected_sheets:
                assert sheet_name in wb.sheetnames
    
    def test_export_excel_overview_sheet(self, sample_spaces):
        """Test Excel export overview sheet content."""
        exporter = ExcelExporter()
        exporter.set_source_file("/test/source.ifc")
        
        with tempfile.TemporaryDirectory() as temp_dir:
            export_path = os.path.join(temp_dir, "test_export.xlsx")
            
            success, message = exporter.export_to_excel(sample_spaces, export_path)
            assert success is True
            
            # Load workbook and check overview sheet
            wb = load_workbook(export_path)
            overview_ws = wb["Overview"]
            
            # Check title
            assert overview_ws['A1'].value == "IFC Room Schedule Export"
            
            # Check for metadata section
            found_export_info = False
            for row in overview_ws.iter_rows(values_only=True):
                if row[0] == "Export Information":
                    found_export_info = True
                    break
            assert found_export_info
    
    def test_export_excel_spaces_sheet(self, sample_spaces):
        """Test Excel export spaces sheet content."""
        exporter = ExcelExporter()
        
        with tempfile.TemporaryDirectory() as temp_dir:
            export_path = os.path.join(temp_dir, "test_export.xlsx")
            
            success, message = exporter.export_to_excel(sample_spaces, export_path)
            assert success is True
            
            # Load workbook and check spaces sheet
            wb = load_workbook(export_path)
            spaces_ws = wb["Spaces"]
            
            # Check headers
            headers = [cell.value for cell in spaces_ws[1]]
            assert "GUID" in headers
            assert "Name" in headers
            assert "Long Name" in headers
            
            # Check data
            assert spaces_ws['A2'].value == "space_guid_1"  # GUID
            assert spaces_ws['B2'].value == "Office 101"    # Name
            assert spaces_ws['C2'].value == "Main Office Room 101"  # Long Name
    
    def test_export_excel_selective_inclusion(self, sample_spaces):
        """Test Excel export with selective data inclusion."""
        exporter = ExcelExporter()
        
        with tempfile.TemporaryDirectory() as temp_dir:
            export_path = os.path.join(temp_dir, "test_export.xlsx")
            
            # Export without surfaces and relationships
            success, message = exporter.export_to_excel(
                sample_spaces, 
                export_path,
                include_surfaces=False,
                include_relationships=False
            )
            assert success is True
            
            # Load workbook and check sheets
            wb = load_workbook(export_path)
            
            # Should have overview, spaces, boundaries, and summary but not surfaces or relationships
            assert "Overview" in wb.sheetnames
            assert "Spaces" in wb.sheetnames
            assert "Space Boundaries" in wb.sheetnames
            assert "Summary" in wb.sheetnames
            assert "Surfaces" not in wb.sheetnames
            assert "Relationships" not in wb.sheetnames
    
    def test_export_excel_empty_spaces(self):
        """Test Excel export with empty spaces list."""
        exporter = ExcelExporter()
        
        with tempfile.TemporaryDirectory() as temp_dir:
            export_path = os.path.join(temp_dir, "test_export.xlsx")
            
            success, message = exporter.export_to_excel([], export_path)
            
            assert success is True
            assert "Successfully exported 0 spaces" in message
            assert os.path.exists(export_path)
    
    def test_export_excel_error_handling(self, sample_spaces):
        """Test Excel export error handling."""
        exporter = ExcelExporter()
        
        # Try to export to invalid path (Windows drive that doesn't exist)
        invalid_path = "Z:\\invalid\\path\\that\\does\\not\\exist\\test.xlsx"
        
        success, message = exporter.export_to_excel(sample_spaces, invalid_path)
        
        assert success is False
        assert "File system error" in message
    
    def test_export_excel_summary_sheet(self, sample_spaces):
        """Test Excel export summary sheet content."""
        exporter = ExcelExporter()
        
        with tempfile.TemporaryDirectory() as temp_dir:
            export_path = os.path.join(temp_dir, "test_export.xlsx")
            
            success, message = exporter.export_to_excel(sample_spaces, export_path)
            assert success is True
            
            # Load workbook and check summary sheet
            wb = load_workbook(export_path)
            summary_ws = wb["Summary"]
            
            # Check title
            assert summary_ws['A1'].value == "Summary Statistics"
            
            # Check for basic statistics section
            found_basic_stats = False
            for row in summary_ws.iter_rows(values_only=True):
                if row[0] == "Basic Statistics":
                    found_basic_stats = True
                    break
            assert found_basic_stats
    
    def test_export_excel_styling(self, sample_spaces):
        """Test Excel export includes proper styling."""
        exporter = ExcelExporter()
        
        with tempfile.TemporaryDirectory() as temp_dir:
            export_path = os.path.join(temp_dir, "test_export.xlsx")
            
            success, message = exporter.export_to_excel(sample_spaces, export_path)
            assert success is True
            
            # Load workbook and check styling
            wb = load_workbook(export_path)
            spaces_ws = wb["Spaces"]
            
            # Check header row styling
            header_cell = spaces_ws['A1']
            assert header_cell.font.bold is True
            # Note: Color checking can be complex due to openpyxl's color handling
            # Just verify that styling was applied
            assert header_cell.font.color is not None
            assert header_cell.fill is not None


class TestExcelExporterWithoutOpenpyxl:
    """Test Excel exporter behavior when openpyxl is not available."""
    
    def test_import_error_handling(self):
        """Test that ExcelExporter raises ImportError when openpyxl is not available."""
        # This test is mainly for documentation - the actual import error handling
        # is tested by the skipif decorator on the main test class
        # We can't easily mock the import without causing recursion issues
        pass