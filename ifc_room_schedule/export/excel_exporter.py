"""
Excel Exporter

Handles exporting room schedule data to Excel format using openpyxl.
"""

import os
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from ..data.space_model import SpaceData


class ExcelExporter:
    """Exports room schedule data to Excel format."""
    
    def __init__(self):
        """Initialize the Excel exporter."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        self.source_file_path: Optional[str] = None
        self.application_version: str = "1.0.0"
        
        # Define styles
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.section_font = Font(bold=True, size=14, color="2F5597")
        self.section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
    
    def set_source_file(self, file_path: str) -> None:
        """Set the source IFC file path."""
        self.source_file_path = file_path
    
    def export_to_excel(self, spaces: List[SpaceData], filename: str,
                       include_surfaces: bool = True,
                       include_boundaries: bool = True,
                       include_relationships: bool = True) -> Tuple[bool, str]:
        """
        Export space data to Excel format.
        
        Args:
            spaces: List of SpaceData objects to export
            filename: Output Excel filename
            include_surfaces: Whether to include surface data
            include_boundaries: Whether to include boundary data
            include_relationships: Whether to include relationship data
            
        Returns:
            Tuple of (success, message)
        """
        try:
            # Validate input
            if not filename:
                return False, "Filename cannot be empty"
            
            if not spaces:
                return False, "No spaces data to export"
            
            # Ensure .xlsx extension
            if not filename.lower().endswith(('.xlsx', '.xls')):
                filename += '.xlsx'
            
            file_path = Path(filename)
            
            # Check write permissions and disk space
            try:
                # Create directory if needed
                file_path.parent.mkdir(parents=True, exist_ok=True)
                
                # Check write permissions
                if file_path.exists():
                    if not os.access(file_path, os.W_OK):
                        return False, f"No write permission for file: {filename}"
                else:
                    if not os.access(file_path.parent, os.W_OK):
                        return False, f"No write permission for directory: {file_path.parent}"
                
                # Check disk space (require at least 5MB for Excel files)
                import shutil
                free_space = shutil.disk_usage(file_path.parent).free
                if free_space < 5 * 1024 * 1024:  # 5MB minimum
                    return False, f"Insufficient disk space. Available: {free_space / (1024*1024):.1f}MB"
                    
            except OSError as e:
                return False, f"File system error: {str(e)}"
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create sheets with error handling
            try:
                self._create_overview_sheet(wb, spaces)
                self._create_spaces_sheet(wb, spaces)
                
                if include_surfaces:
                    self._create_surfaces_sheet(wb, spaces)
                
                if include_boundaries:
                    self._create_boundaries_sheet(wb, spaces)
                
                if include_relationships:
                    self._create_relationships_sheet(wb, spaces)
                
                self._create_summary_sheet(wb, spaces)
                
            except Exception as e:
                return False, f"Error creating Excel sheets: {str(e)}"
            
            # Save workbook with atomic operation
            temp_filename = str(file_path) + '.tmp'
            try:
                wb.save(temp_filename)
                
                # Atomic rename
                if os.name == 'nt':  # Windows
                    if file_path.exists():
                        os.remove(file_path)
                os.rename(temp_filename, filename)
                
                return True, f"Successfully exported {len(spaces)} spaces to {Path(filename).name}"
                
            except Exception as e:
                # Clean up temp file
                if os.path.exists(temp_filename):
                    try:
                        os.remove(temp_filename)
                    except:
                        pass
                return False, f"Error saving Excel file: {str(e)}"
            
        except PermissionError as e:
            return False, f"Permission denied: {str(e)}"
        except OSError as e:
            return False, f"OS error: {str(e)}"
        except Exception as e:
            return False, f"Excel export failed: {str(e)}"
    
    def _create_overview_sheet(self, wb: Workbook, spaces: List[SpaceData]) -> None:
        """Create overview sheet with metadata and summary."""
        ws = wb.create_sheet("Overview", 0)
        
        # Title
        ws['A1'] = "IFC Room Schedule Export"
        ws['A1'].font = Font(bold=True, size=16, color="2F5597")
        ws.merge_cells('A1:D1')
        
        # Metadata section
        row = 3
        ws[f'A{row}'] = "Export Information"
        ws[f'A{row}'].font = self.section_font
        ws[f'A{row}'].fill = self.section_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        metadata = [
            ("Export Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Application Version:", self.application_version),
            ("Total Spaces:", len(spaces)),
            ("Processed Spaces:", sum(1 for space in spaces if space.processed))
        ]
        
        if self.source_file_path:
            metadata.insert(2, ("Source File:", Path(self.source_file_path).name))
        
        for label, value in metadata:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        # Quick statistics
        row += 1
        ws[f'A{row}'] = "Quick Statistics"
        ws[f'A{row}'].font = self.section_font
        ws[f'A{row}'].fill = self.section_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        total_surfaces = sum(len(space.surfaces) for space in spaces)
        total_boundaries = sum(len(space.space_boundaries) for space in spaces)
        total_relationships = sum(len(space.relationships) for space in spaces)
        total_surface_area = sum(space.get_total_surface_area() for space in spaces)
        total_boundary_area = sum(space.get_total_boundary_area() for space in spaces)
        
        stats = [
            ("Total Surfaces:", total_surfaces),
            ("Total Boundaries:", total_boundaries),
            ("Total Relationships:", total_relationships),
            ("Total Surface Area (m²):", round(total_surface_area, 2)),
            ("Total Boundary Area (m²):", round(total_boundary_area, 2))
        ]
        
        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _create_spaces_sheet(self, wb: Workbook, spaces: List[SpaceData]) -> None:
        """Create spaces data sheet."""
        ws = wb.create_sheet("Spaces")
        
        # Headers
        headers = [
            'GUID', 'Name', 'Long Name', 'Description', 'Object Type',
            'Zone Category', 'Number', 'Elevation', 'Processed',
            'Height', 'Finish Floor Height', 'Finish Ceiling Height',
            'Total Surface Area (m²)', 'Total Boundary Area (m²)', 'User Description'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Write data
        for row, space in enumerate(spaces, 2):
            quantities = space.quantities or {}
            user_desc = space.user_descriptions.get('space', '') if space.user_descriptions else ''
            
            data = [
                space.guid,
                space.name or '',
                space.long_name or '',
                space.description or '',
                space.object_type or '',
                space.zone_category or '',
                space.number or '',
                space.elevation or 0.0,
                'Yes' if space.processed else 'No',
                quantities.get('Height', ''),
                quantities.get('FinishFloorHeight', ''),
                quantities.get('FinishCeilingHeight', ''),
                round(space.get_total_surface_area(), 2),
                round(space.get_total_boundary_area(), 2),
                user_desc
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _create_surfaces_sheet(self, wb: Workbook, spaces: List[SpaceData]) -> None:
        """Create surfaces data sheet."""
        ws = wb.create_sheet("Surfaces")
        
        # Headers
        headers = [
            'Space GUID', 'Space Name', 'Surface ID', 'Surface Type',
            'Area (m²)', 'Material', 'IFC Type', 'User Description'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Write data
        row = 2
        for space in spaces:
            for surface in space.surfaces:
                data = [
                    space.guid,
                    space.name or '',
                    surface.id,
                    surface.type or '',
                    round(surface.area, 2) if surface.area else 0.0,
                    surface.material or '',
                    surface.ifc_type or '',
                    surface.user_description or ''
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = self.border
                row += 1
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _create_boundaries_sheet(self, wb: Workbook, spaces: List[SpaceData]) -> None:
        """Create space boundaries data sheet."""
        ws = wb.create_sheet("Space Boundaries")
        
        # Headers
        headers = [
            'Space GUID', 'Space Name', 'Boundary GUID', 'Boundary Name',
            'Physical/Virtual', 'Internal/External', 'Surface Type', 'Orientation',
            'Area (m²)', 'Related Element GUID', 'Related Element Name', 'Related Element Type',
            'Adjacent Space GUID', 'Adjacent Space Name', 'Boundary Level', 'Display Label',
            'User Description'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Write data
        row = 2
        for space in spaces:
            for boundary in space.space_boundaries:
                data = [
                    space.guid,
                    space.name or '',
                    boundary.guid,
                    boundary.name or '',
                    boundary.physical_or_virtual_boundary or '',
                    boundary.internal_or_external_boundary or '',
                    boundary.boundary_surface_type or '',
                    boundary.boundary_orientation or '',
                    round(boundary.calculated_area, 2) if boundary.calculated_area else 0.0,
                    boundary.related_building_element_guid or '',
                    boundary.related_building_element_name or '',
                    boundary.related_building_element_type or '',
                    boundary.adjacent_space_guid or '',
                    boundary.adjacent_space_name or '',
                    boundary.boundary_level or 1,
                    boundary.display_label or '',
                    boundary.user_description or ''
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = self.border
                row += 1
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _create_relationships_sheet(self, wb: Workbook, spaces: List[SpaceData]) -> None:
        """Create relationships data sheet."""
        ws = wb.create_sheet("Relationships")
        
        # Headers
        headers = [
            'Space GUID', 'Space Name', 'Related Entity GUID', 'Related Entity Name',
            'Related Entity Description', 'Relationship Type', 'IFC Relationship Type'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Write data
        row = 2
        for space in spaces:
            for relationship in space.relationships:
                data = [
                    space.guid,
                    space.name or '',
                    relationship.related_entity_guid,
                    relationship.related_entity_name or '',
                    relationship.related_entity_description or '',
                    relationship.relationship_type or '',
                    relationship.ifc_relationship_type or ''
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = self.border
                row += 1
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _create_summary_sheet(self, wb: Workbook, spaces: List[SpaceData]) -> None:
        """Create summary statistics sheet."""
        ws = wb.create_sheet("Summary")
        
        # Title
        ws['A1'] = "Summary Statistics"
        ws['A1'].font = Font(bold=True, size=16, color="2F5597")
        ws.merge_cells('A1:B1')
        
        # Calculate summary statistics
        total_spaces = len(spaces)
        processed_spaces = sum(1 for space in spaces if space.processed)
        total_surfaces = sum(len(space.surfaces) for space in spaces)
        total_boundaries = sum(len(space.space_boundaries) for space in spaces)
        total_relationships = sum(len(space.relationships) for space in spaces)
        total_surface_area = sum(space.get_total_surface_area() for space in spaces)
        total_boundary_area = sum(space.get_total_boundary_area() for space in spaces)
        
        # Basic statistics
        row = 3
        ws[f'A{row}'] = "Basic Statistics"
        ws[f'A{row}'].font = self.section_font
        ws[f'A{row}'].fill = self.section_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        basic_stats = [
            ("Total Spaces", total_spaces),
            ("Processed Spaces", processed_spaces),
            ("Total Surfaces", total_surfaces),
            ("Total Boundaries", total_boundaries),
            ("Total Relationships", total_relationships),
            ("Total Surface Area (m²)", round(total_surface_area, 2)),
            ("Total Boundary Area (m²)", round(total_boundary_area, 2))
        ]
        
        for label, value in basic_stats:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        # Surface areas by type
        row += 1
        ws[f'A{row}'] = "Surface Areas by Type (m²)"
        ws[f'A{row}'].font = self.section_font
        ws[f'A{row}'].fill = self.section_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        surface_area_by_type = {}
        for space in spaces:
            space_areas = space.get_surface_area_by_type()
            for surface_type, area in space_areas.items():
                if surface_type not in surface_area_by_type:
                    surface_area_by_type[surface_type] = 0.0
                surface_area_by_type[surface_type] += area
        
        for surface_type, area in surface_area_by_type.items():
            ws[f'A{row}'] = surface_type
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = round(area, 2)
            row += 1
        
        # Boundary areas by type
        row += 1
        ws[f'A{row}'] = "Boundary Areas by Type (m²)"
        ws[f'A{row}'].font = self.section_font
        ws[f'A{row}'].fill = self.section_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        boundary_area_by_type = {}
        for space in spaces:
            space_boundary_areas = space.get_boundary_area_by_type()
            for boundary_type, area in space_boundary_areas.items():
                if boundary_type not in boundary_area_by_type:
                    boundary_area_by_type[boundary_type] = 0.0
                boundary_area_by_type[boundary_type] += area
        
        for boundary_type, area in boundary_area_by_type.items():
            ws[f'A{row}'] = boundary_type
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = round(area, 2)
            row += 1
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _auto_adjust_columns(self, ws) -> None:
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set minimum width and maximum width
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width